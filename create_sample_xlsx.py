"""
create_sample_xlsx.py - Generate a realistic sample security questionnaire xlsx
for demo and testing purposes. Run once: python create_sample_xlsx.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_sample():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Access Control ----
    ws1 = wb.active
    ws1.title = "Access Control"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, color="FFFFFF", size=12)

    # Title row
    ws1["A1"] = "Security Questionnaire v2.1"
    ws1["A1"].font = Font(bold=True, size=14)

    ws1["A2"] = "Instructions: Please complete all fields. Mark N/A where not applicable."
    ws1["A2"].font = Font(italic=True)

    # Section header
    ws1["A4"] = "ACCESS CONTROL"
    ws1["A4"].font = header_font

    # Column headers
    ws1["A5"] = "Question ID"
    ws1["B5"] = "Question"
    ws1["C5"] = "Response"
    ws1["D5"] = "Comments / Evidence"
    for col in ["A5", "B5", "C5", "D5"]:
        ws1[col].font = header_font_white
        ws1[col].fill = header_fill

    questions_ac = [
        ("AC-01", "Does your organization have a formal access control policy?", "Yes / No / N/A", ""),
        ("AC-02", "Do you enforce multi-factor authentication (MFA) for all administrative accounts?", "Yes / No / N/A", ""),
        ("AC-03", "Is privileged access management (PAM) in place for all critical systems?", "Yes / No / N/A", ""),
        ("AC-04", "Describe your process for provisioning and deprovisioning user access.", "Free text", ""),
        ("AC-05", "How often are user access rights reviewed?", "A) Monthly\nB) Quarterly\nC) Annually\nD) Never", ""),
        ("AC-06", "Are all default passwords changed upon system deployment?", "Yes / No / N/A", ""),
        ("AC-07", "Do you maintain an inventory of all user accounts with privileged access?", "Yes / No / N/A", ""),
    ]

    for i, (qid, text, response, comment) in enumerate(questions_ac, start=6):
        ws1[f"A{i}"] = qid
        ws1[f"B{i}"] = text
        ws1[f"C{i}"] = response
        ws1[f"D{i}"] = comment

    # Set column widths
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 55
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 30

    # ---- Sheet 2: Encryption ----
    ws2 = wb.create_sheet("Encryption")

    ws2["A1"] = "ENCRYPTION & KEY MANAGEMENT"
    ws2["A1"].font = header_font

    ws2["A2"] = "Question ID"
    ws2["B2"] = "Question"
    ws2["C2"] = "Response"

    questions_enc = [
        ("EKM-01", "Is data encrypted at rest using AES-256 or equivalent?", "Yes / No / N/A"),
        ("EKM-02", "Is all data in transit protected using TLS 1.2 or higher?", "Yes / No / N/A"),
        ("EKM-03", "Explain your key management practices and rotation schedule.", "Free text"),
        ("EKM-04", "Are encryption keys stored separately from encrypted data?", "Yes / No / N/A"),
        ("EKM-05", "Do you use a Hardware Security Module (HSM) for key storage?", "Yes / No / N/A"),
    ]

    for i, (qid, text, response) in enumerate(questions_enc, start=3):
        ws2[f"A{i}"] = qid
        ws2[f"B{i}"] = text
        ws2[f"C{i}"] = response

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 30

    # ---- Sheet 3: Incident Response ----
    ws3 = wb.create_sheet("Incident Response")

    ws3["A1"] = "Version: 2.1"  # Test version detection

    ws3["A3"] = "INCIDENT RESPONSE"
    ws3["A3"].font = header_font

    ws3["A4"] = "Question"
    ws3["B4"] = "Response Options"

    questions_ir = [
        ("Do you have a documented incident response plan?", "Yes / No / N/A"),
        ("How quickly does your organization respond to a confirmed security incident?", "A) Within 1 hour\nB) Within 4 hours\nC) Within 24 hours\nD) Within 72 hours"),
        ("Describe your process for notifying customers of a data breach.", "Free text"),
        ("Is your incident response plan tested at least annually?", "Yes / No / N/A"),
        ("Are incident response roles and responsibilities clearly defined?", "Yes / No / N/A"),
    ]

    for i, (text, response) in enumerate(questions_ir, start=5):
        ws3[f"A{i}"] = text
        ws3[f"B{i}"] = response

    ws3.column_dimensions["A"].width = 60
    ws3.column_dimensions["B"].width = 35

    wb.save("data/sample_questionnaire.xlsx")
    print("Created data/sample_questionnaire.xlsx")


if __name__ == "__main__":
    create_sample()
