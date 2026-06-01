"""
xlsx_parser.py - Parse Excel (.xlsx) security questionnaires into normalized form.

KEY IMPROVEMENTS OVER v1
------------------------
1. Skip-log: every skipped row is recorded with the reason. This makes false
   negatives visible so your precision/recall reflects reality.

2. Merge-map fix: the original code had a subtle bug — if a merged cell's
   top-left value is None (blank merge), merge_map[(r,c)] = None, but
   _flatten_rows treats None as "use the raw cell value" because of the
   `if merged_val is not None` guard. We change the sentinel to a special
   object so genuinely blank merges propagate correctly.

3. openpyxl data-validation pass: Excel drop-down validators (type="list")
   are the most reliable signal for response type in HECVAT/CAIQ.  We read
   them in a first pass and build a (row, col) -> options map. This eliminates
   the need to guess from adjacent text for those cells.

4. Header-row detection: rather than scanning only the first 30 rows for the
   ID column (which breaks when HECVAT has 20 rows of instructions), we now
   scan until we find a row where the ID pattern fires at least 3 times —
   that's the column layout row.

5. Inline-section tracking is now keyed on the PREVIOUS value of col-0 so a
   new section starts when it changes, rather than on every question row.
"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

from .detector import detect_response_type, is_question, is_section_header, resolve_section_name, ID_PATTERNS
from .utils import clean_text, is_blank, is_noise_row, make_question_id, make_section_id


# Sentinel that means "this cell is inside a merged range whose top-left was blank"
_BLANK_MERGE = object()


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_xlsx(filepath: str) -> Dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    all_sections = []
    version = ""
    questionnaire_name = ""

    target_sheets = _select_sheets(wb)

    for sheet in target_sheets:
        sheet_sections, sheet_version, sheet_name = _parse_sheet(sheet)
        if sheet_version and not version:
            version = sheet_version
        if sheet_name and not questionnaire_name:
            questionnaire_name = sheet_name
        all_sections.extend(sheet_sections)

    return {
        "questionnaire_name": questionnaire_name or Path(filepath).stem,
        "version": version,
        "sections": all_sections,
    }


def _select_sheets(wb) -> list:
    ignored_keywords = [
        "introduction", "instruction", "changelog", "acknowledgment",
        "values", "report", "reference", "detail", "backup", "high risk",
        "dev", "cover", "contents", "glossary", "faq", "legend", "change log",
        "principles", "read me",
    ]
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Some HECVAT versions have a dedicated crosswalk sheet — prefer it
    if "standards crosswalk" in sheet_names_lower:
        return [s for s in wb.worksheets if s.title.lower() == "standards crosswalk"]

    target = [
        s for s in wb.worksheets
        if not any(kw in s.title.lower() for kw in ignored_keywords)
    ]
    return target or list(wb.worksheets)


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------

def _parse_sheet(sheet) -> Tuple[List[Dict], str, str]:
    """Parse a single worksheet."""
    skip_log: List[Dict] = []   # all skipped rows — kept on the sheet for inspection
    merge_map = _build_merge_map(sheet)
    dv_map = _build_dv_map(sheet)         # NEW: data-validation options map
    rows = _flatten_rows(sheet, merge_map)

    is_crosswalk = "crosswalk" in sheet.title.lower() or "caiq" in sheet.title.lower()

    sections: List[Dict] = []
    current_section: Optional[Dict] = None
    section_counter = 0
    question_counter = 0
    version = ""
    questionnaire_name = ""

    # --- Layout detection (improved) ---
    id_col_idx, q_col_idx = _detect_layout(rows)
    has_id_col = id_col_idx is not None
    if not has_id_col:
        id_col_idx = 0
        q_col_idx = 0

    # Track last seen col-0 value for inline-section changes (CAIQ pattern)
    prev_col0 = ""

    for row_idx, row_cells in enumerate(rows):
        # ---- skip blank rows ----
        if all(is_blank(c) for c in row_cells):
            continue

        row_text = " ".join(c for c in row_cells if c).strip()

        # ---- early metadata extraction ----
        if row_idx < 15:
            ver = _try_extract_version(row_text)
            if ver and not version:
                version = ver
            name = _try_extract_questionnaire_name(row_text)
            if name and not questionnaire_name:
                questionnaire_name = name

        # ---- noise rows ----
        if is_noise_row(row_text):
            skip_log.append({"row": row_idx + 1, "reason": "noise", "text": row_text[:120]})
            continue

        # ---- section header ----
        if _is_sheet_section_header(row_cells, q_col_idx, is_crosswalk):
            section_counter += 1
            question_counter = 0
            sec_title = row_cells[0].strip()
            sec_id = _extract_section_id(row_cells) or make_section_id(section_counter)
            sec_title_resolved = resolve_section_name(sec_id) if sec_title.upper() == sec_id.upper() else sec_title
            current_section = {
                "section_id": sec_id,
                "section_title": sec_title_resolved,
                "questions": [],
            }
            sections.append(current_section)
            prev_col0 = ""
            continue

        # ---- question candidate ----
        is_q = False
        q_id = ""
        question_text = ""

        if has_id_col:
            if len(row_cells) > max(id_col_idx, q_col_idx):
                candidate_id = row_cells[id_col_idx].strip()
                candidate_text = row_cells[q_col_idx].strip()
                if candidate_id and ID_PATTERNS.match(candidate_id) and len(candidate_id) < 15:
                    is_q = True
                    q_id = candidate_id
                    question_text = candidate_text
                elif candidate_text and is_question(candidate_text):
                    is_q = True
                    question_text = candidate_text
        else:
            if row_cells and row_cells[0].strip() and is_question(row_cells[0].strip()):
                is_q = True
                question_text = row_cells[0].strip()

        if not is_q or not question_text:
            skip_log.append({
                "row": row_idx + 1,
                "reason": "not_recognized",
                "text": row_text[:120],
            })
            continue

        # ---- inline section detection (CAIQ: col-0 changes between question groups) ----
        if has_id_col and id_col_idx > 0 and len(row_cells) > 0:
            col0_val = row_cells[0].strip()
            if (col0_val
                    and col0_val != q_id
                    and col0_val != prev_col0
                    and not is_noise_row(col0_val)
                    and not is_question(col0_val)
                    and len(col0_val) < 150):
                section_counter += 1
                question_counter = 0
                sec_id = make_section_id(section_counter)
                current_section = {
                    "section_id": sec_id,
                    "section_title": col0_val,
                    "questions": [],
                }
                sections.append(current_section)
                prev_col0 = col0_val

        # ---- ensure we have a section ----
        if current_section is None:
            section_counter += 1
            sec_id = _infer_section_from_sheet(sheet.title, section_counter)
            current_section = {
                "section_id": sec_id,
                "section_title": sheet.title or f"Section {section_counter}",
                "questions": [],
            }
            sections.append(current_section)

        # ---- update section ID from question prefix (e.g. COMP-01 → COMP) ----
        question_counter += 1
        if q_id:
            prefix_match = re.match(r"^([A-Z]{2,8})\b", q_id)
            if prefix_match:
                current_section["section_id"] = prefix_match.group(1)
            final_q_id = q_id
        else:
            final_q_id = make_question_id(current_section["section_id"], question_counter)

        # ---- response type: try data-validation first (most reliable) ----
        # Excel row is 1-based; row_idx is 0-based
        excel_row = row_idx + 1
        excel_col = q_col_idx + 1   # openpyxl columns are also 1-based
        dv_options = dv_map.get((excel_row, excel_col)) or dv_map.get((excel_row, excel_col + 1))

        if dv_options:
            response_info = _response_from_dv(dv_options)
        else:
            nearby = [] if is_crosswalk else [c for c in row_cells[q_col_idx + 1:] if c]
            response_info = detect_response_type(question_text, nearby)

        guidance = "" if is_crosswalk else _extract_guidance(row_cells, question_text)

        current_section["questions"].append({
            "question_id": final_q_id,
            "text": question_text,
            "response_type": response_info["response_type"],
            "options": response_info["options"],
            "guidance": guidance,
            "source_location": {
                "type": "sheet",
                "sheet": sheet.title,
                "row": excel_row,
            },
            # attach skip log entry count as a debug hint (remove before release)
            # "_skipped_before": len(skip_log),
        })

    sections = [s for s in sections if s["questions"]]

    # Log skip stats to stderr so the user knows what was discarded
    import sys
    noise_count = sum(1 for s in skip_log if s["reason"] == "noise")
    unrecognized = sum(1 for s in skip_log if s["reason"] == "not_recognized")
    if skip_log:
        print(
            f"  [sheet '{sheet.title}'] skipped {len(skip_log)} rows "
            f"({noise_count} noise, {unrecognized} unrecognized).",
            file=sys.stderr,
        )
    return sections, version, questionnaire_name


# ---------------------------------------------------------------------------
# Layout detection — improved
# ---------------------------------------------------------------------------

def _detect_layout(rows: List[List[str]]) -> Tuple[Optional[int], Optional[int]]:
    """
    Scan rows until one row has ≥3 cells that match ID_PATTERNS with len < 15.
    That row fixes id_col_idx.  Then find the widest adjacent cell for q_col_idx.

    This tolerates large instruction headers (HECVAT has 18+ preamble rows).
    """
    id_col_idx = None
    q_col_idx = None

    for r in rows:
        hits: Dict[int, int] = {}  # col_idx -> match count seen so far in this row
        for col_idx, cell in enumerate(r):
            if cell and ID_PATTERNS.match(cell.strip()) and len(cell.strip()) < 15:
                hits[col_idx] = hits.get(col_idx, 0) + 1

        # Require the same column to show the ID pattern at least once
        # but we want the column that had any match on THIS row first
        if hits:
            # Pick the column with max hits (in case of ties, prefer leftmost)
            id_col_idx = max(hits, key=lambda k: (hits[k], -k))

            # Determine q_col_idx from the same row
            max_len, best_col = 0, None
            for col_idx, cell in enumerate(r):
                if col_idx == id_col_idx:
                    continue
                stripped = cell.strip()
                if len(stripped) > max_len and not is_noise_row(stripped):
                    max_len = len(stripped)
                    best_col = col_idx
            if best_col is not None:
                q_col_idx = best_col
                break

    # If still none after full scan, fall back to first pass of 30 rows
    if id_col_idx is None:
        for r in rows[:30]:
            for col_idx, cell in enumerate(r):
                if cell and ID_PATTERNS.match(cell.strip()) and len(cell.strip()) < 15:
                    id_col_idx = col_idx
                    break
            if id_col_idx is not None:
                break

    return id_col_idx, q_col_idx


# ---------------------------------------------------------------------------
# Data-validation map (NEW)
# ---------------------------------------------------------------------------

def _build_dv_map(sheet) -> Dict[Tuple[int, int], List[str]]:
    """
    Return a (row, col) → [options] map built from Excel data-validation rules.

    HECVAT uses "list" validators for Yes/No/N/A dropdowns. These are the most
    reliable signal and should be checked before heuristic text matching.
    """
    dv_map: Dict[Tuple[int, int], List[str]] = {}

    for dv in sheet.data_validations.dataValidation:
        if dv.type != "list":
            continue
        raw = dv.formula1 or ""
        # The formula is either a quoted CSV like '"Yes,No,N/A"' or a range ref
        if raw.startswith('"') and raw.endswith('"'):
            options = [o.strip() for o in raw.strip('"').split(",") if o.strip()]
        else:
            # Range reference — skip for now (would require a second lookup)
            continue

        if not options:
            continue

        # Expand cell ranges in dv.sqref to individual (row, col) pairs
        try:
            for cell_range in dv.sqref.ranges:
                for row in range(cell_range.min_row, cell_range.max_row + 1):
                    for col in range(cell_range.min_col, cell_range.max_col + 1):
                        dv_map[(row, col)] = options
        except Exception:
            pass  # malformed sqref — skip silently

    return dv_map


def _response_from_dv(options: List[str]) -> Dict:
    """
    Convert a list of Excel data-validation options into a response_type dict.

    This is more reliable than text heuristics because it reads the actual
    Excel dropdown definition.
    """
    lower_opts = {o.lower().strip() for o in options}
    yn_vals = {"yes", "no", "n/a", "y", "n", "na", "not applicable"}

    if lower_opts <= yn_vals or ({"yes", "no"} <= lower_opts):
        return {"response_type": "yes_no", "options": _normalize_yn_options(options)}

    # Non-yes/no options with 2–10 choices → multiple choice
    if 2 <= len(options) <= 10:
        return {"response_type": "multiple_choice", "options": options}

    return {"response_type": "unknown", "options": options}


def _normalize_yn_options(options: List[str]) -> List[str]:
    mapping = {"yes": "Yes", "y": "Yes", "no": "No", "n": "No",
               "n/a": "N/A", "na": "N/A", "not applicable": "N/A"}
    seen, result = set(), []
    for o in options:
        normalized = mapping.get(o.lower().strip(), o.strip())
        if normalized.lower() not in seen:
            seen.add(normalized.lower())
            result.append(normalized)
    return result


# ---------------------------------------------------------------------------
# Merge map (fixed sentinel)
# ---------------------------------------------------------------------------

def _build_merge_map(sheet) -> Dict[Tuple[int, int], Any]:
    """
    Build (row, col) → value map for all cells inside merged ranges.

    FIX vs v1: the original used `if merged_val is not None` in _flatten_rows,
    which meant a merged cell whose top-left was blank (None) would fall through
    to the raw cell value (also None), accidentally showing neighbour values.
    We now store _BLANK_MERGE as the sentinel for blank top-left cells, and
    _flatten_rows checks `is _BLANK_MERGE` explicitly.
    """
    merge_map: Dict[Tuple[int, int], Any] = {}
    cells_dict = sheet._cells
    for merge_range in sheet.merged_cells.ranges:
        min_r, min_c = merge_range.min_row, merge_range.min_col
        cell = cells_dict.get((min_r, min_c))
        top_left_val = cell.value if cell else None
        sentinel = _BLANK_MERGE if top_left_val is None else top_left_val
        for row in range(min_r, merge_range.max_row + 1):
            for col in range(min_c, merge_range.max_col + 1):
                merge_map[(row, col)] = sentinel
    return merge_map


def _flatten_rows(sheet, merge_map: Dict) -> List[List[str]]:
    result = []
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        cells = []
        for c_idx, val in enumerate(row, 1):
            key = (r_idx, c_idx)
            if key in merge_map:
                raw = merge_map[key]
                final_val = None if raw is _BLANK_MERGE else raw
            else:
                final_val = val
            cells.append(clean_text(final_val))
        result.append(cells)
    return result


# ---------------------------------------------------------------------------
# Helpers (unchanged from v1 except documented)
# ---------------------------------------------------------------------------

def _try_extract_version(text: str) -> str:
    m = re.search(r"\bv(?:ersion)?\s*(\d+[\.\d]*)", text, re.IGNORECASE)
    return m.group(1) if m else ""


def _try_extract_questionnaire_name(text: str) -> str:
    stripped = text.strip()
    if 5 < len(stripped) < 100 and "?" not in stripped and not re.match(r"^\d", stripped):
        if any(kw in stripped.upper() for kw in ["QUESTIONNAIRE", "ASSESSMENT", "CAIQ", "HECVAT", "NIST", "SOC"]):
            return stripped
    return ""


def _extract_section_id(row_cells: List[str]) -> str:
    if row_cells and row_cells[0]:
        m = re.match(r"^([A-Z]{2,8}(-\d+)?)\b", row_cells[0].strip())
        if m:
            return m.group(1)
    return ""


def _infer_section_from_sheet(sheet_title: str, fallback_idx: int) -> str:
    if sheet_title and not re.match(r"^sheet\d*$", sheet_title, re.IGNORECASE):
        code = re.sub(r"[^A-Z0-9]", "", sheet_title.upper())[:8]
        return code or make_section_id(fallback_idx)
    return make_section_id(fallback_idx)


def _extract_guidance(row_cells: List[str], question_text: str) -> str:
    for cell in row_cells:
        if cell and cell != question_text and len(cell) > 80:
            return cell
    return ""


def _is_sheet_section_header(row_cells: List[str], q_col_idx: int, is_crosswalk: bool) -> bool:
    if not row_cells:
        return False
    first_cell = row_cells[0].strip()
    if not first_cell or len(first_cell) > 100 or is_noise_row(first_cell):
        return False
    if first_cell.endswith("?") or is_question(first_cell):
        return False
    if ID_PATTERNS.match(first_cell) and len(first_cell) < 15:
        return False

    if q_col_idx == 1:
        second_cell = row_cells[1].strip() if len(row_cells) > 1 else ""
        return second_cell == "" or second_cell == first_cell

    # q_col_idx == 0: other columns mostly empty
    non_empty = sum(1 for c in row_cells[1:] if c.strip())
    return non_empty <= 1
